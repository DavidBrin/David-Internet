#imports

import sys
sys.path.append(r"..\spikeparam")

print(sys.path)

import spikeparam
from spikeparam.patch.fit import Spike
from spikeparam.patch.fit import SpikeGroup

from neurodsp import spectral

from scipy import signal
import scipy

import h5py
from tqdm import tqdm

import numpy as np
import pandas as pd

from neurodsp import filt
from neurodsp.timefrequency import amp_by_time, phase_by_time
from neurodsp.plts import plot_time_series, plot_instantaneous_measure
from neurodsp.plts.time_series import plot_bursts
from neurodsp.burst import detect_bursts_dual_threshold, compute_burst_stats

from scipy.signal import sosfiltfilt, butter

from scipy.signal import find_peaks
from scipy.optimize import curve_fit
from scipy.stats import pearsonr
import statsmodels.api as sm
from sklearn.metrics.pairwise import cosine_similarity

import matplotlib.pyplot as plt
import seaborn as sns
import scipy.stats as stats
import scipy.spatial as sp_spatial

import os

from fooof import FOOOF

sns.set(rc={'figure.figsize':(12,9)})
sns.set_style('whitegrid')
sns.set_style("whitegrid", {'axes.grid' : False})

import IProgress

import openpyxl
import glob

import pickle

# Set interactive mode to display plots inline
#plt.ion()

# Set the figure DPI for high-resolution plotting
plt.rcParams['figure.dpi'] = 300  # Adjust DPI value as needed for your display

# Functions for saving and loading dataframes to/from pickle files

def save_dataframe_to_pickle(dataframe, file_path):
    """
    Function to save a DataFrame as a pickle file.
    
    Args:
    - dataframe (pd.DataFrame): DataFrame to be saved.
    - file_path (str): Path to save the pickle file.
    """
    dataframe.to_pickle(file_path)

def load_dataframe_from_pickle(file_path):
    """
    Function to extract a DataFrame from a pickle file.
    
    Args:
    - file_path (str): Path to the pickle file.
    
    Returns:
    - dataframe (pd.DataFrame): Loaded DataFrame.
    """
    dataframe = pd.read_pickle(file_path)
    return dataframe

def save_dict_to_pickle(dictionary, filepath):
    """
    Save a dictionary to a pickle file.

    Parameters:
        dictionary (dict): The dictionary to save.
        filepath (str): The path to the pickle file.
    """
    with open(filepath, 'wb') as f:
        pickle.dump(dictionary, f)
    print(f"Dictionary saved to {filepath}")


def load_dict_from_pickle(filepath):
    """
    Load a dictionary from a pickle file.

    Parameters:
        filepath (str): The path to the pickle file.

    Returns:
        dict: The loaded dictionary.
    """
    with open(filepath, 'rb') as f:
        dictionary = pickle.load(f)
    print(f"Dictionary loaded from {filepath}")
    return dictionary

# Function to visualize sweep patch data
# no metadata, just the time series data
def extract_data(file_path, plot_data=False):
    """
    Extracts data from an HDF5 file and optionally plots the data.

    This function opens an HDF5 file, extracts data from each sweep in the 'acquisition' group, 
    and stores the data in a list of NumPy arrays. If the plot_data flag is set to True, the 
    function also plots the data. It handles both 1D and 2D data, but prints a message if the 
    data has more than 2 dimensions.

    Parameters:
    - file_path (str): The path to the HDF5 file.
    - plot_data (bool): A flag indicating whether to plot the extracted data. Default is False.

    Returns:
    - list: A list of NumPy arrays containing the extracted data for each sweep.
    """
    # Open the HDF5 file
    with h5py.File(file_path, 'r') as f:
        # Initialize an empty list to store data arrays
        data = []

        # Iterate through keys in the 'acquisition' group
        for sweep_key in f['acquisition'].keys():
            dataset = f['acquisition'][sweep_key]['data'] 
            # Convert the dataset data into a NumPy array and append to the list
            data.append(np.array(dataset))

        # Plot the data
        if(plot_data):
            if all(d.ndim == 1 for d in data):
                for d in data:
                    plt.plot(d)
                plt.xlabel('time (ms)')
                plt.ylabel('mV')
                plt.title('1D Dataset Visualization')
                plt.show()
            elif all(d.ndim == 2 for d in data):
                for d in data:
                    plt.imshow(d, cmap='viridis')
                    plt.colorbar()
                    plt.xlabel('X-axis')
                    plt.ylabel('Y-axis')
                    plt.title('2D Dataset Visualization')
                    plt.show()
            else:
                print("Cannot visualize data with more than 2 dimensions.")

        return data


def update_columns_at_index(df, file_path, index):
    """
    Function to update columns in the DataFrame at a specific index with values from Excel metadata.
    
    Args:
    - df (pd.DataFrame): DataFrame to update.
    - file_path (str): Path to the Excel file containing metadata.
    - index (int): Index at which to update the columns.
    
    Returns:
    - df (pd.DataFrame): Updated DataFrame.
    """
    # Extract filename from file_path
    filename = os.path.splitext(os.path.basename(file_path))[0]
    
    # Load the workbook
    wb = openpyxl.load_workbook(file_path)
    # Select the active worksheet
    ws = wb.active
    
    # Extract the value from the first cell of the row
    first_cell_value = ws.cell(row=index, column=1).value
    
    # Check if filename matches the string in the first cell of the row
    if filename != first_cell_value:
        return False
    
    # Extract column names from the first row of the Excel sheet
    column_names = [cell.value for cell in ws[1]]
    
    # Extract data from the specified row of the Excel sheet
    data_row = next(ws.iter_rows(min_row=index, max_row=index, values_only=True))
    
    # Update DataFrame columns at the specified index with new data
    for column_name, value in zip(column_names, data_row):
        df[column_name] = value
    
    return True


def monkey_df(filepaths, ind_start):
    """
    Creates a consolidated DataFrame containing spike data features from multiple HDF5 files.

    This function iterates through each HDF5 file in the provided list of file paths, extracts 
    the spike data for each sweep in the 'acquisition' group, fits a spike detection algorithm 
    to the data, and collects the resulting features into a DataFrame. The DataFrames from each 
    file are then concatenated into a single DataFrame, which is returned. Each sweep and file 
    is labeled accordingly.

    Parameters:
    - filepaths (list of str): List of file paths to the HDF5 files.
    - ind_start (int): Starting index for labeling the files.

    Returns:
    - pandas.DataFrame: A consolidated DataFrame containing spike data features from all the files.
    """
    
    # Initialize the final DataFrame
    super_mega_df = pd.DataFrame()
    index = ind_start

    # Iterate through each file in the file paths
    for file in filepaths:
        print(file)
        data = extract_data(file, plot_data=False)  # Extract data from the file (numpy array of all sweeps)
        spike_dir = {}  # Initialize dictionary to store spike objects for each sweep

        # Open the HDF5 file and process each sweep
        with h5py.File(file, 'r') as f:
            i = 0
            for sweep_key in f['acquisition'].keys():
                if i < len(data):
                    try:
                        # Initialize and fit the Spike object to the data
                        sweep_key_obj = Spike(thresh_amp=0, window_length=(5., 5.), smooth_frac=.01)
                        sweep_key_obj.fit(data[i], 20000, n_jobs=-1, progress=tqdm)
                        
                        # If spikes are detected, add the object to the dictionary
                        if sweep_key_obj.n_spikes is not None:
                            spike_dir[sweep_key] = sweep_key_obj
                    except ValueError as e:
                        print(f"Fitting failed for sweep {sweep_key}: {e}")
                i += 1

        # Create a DataFrame for the current file
        mega_df = pd.DataFrame()
        for sweep_key, spike_obj in spike_dir.items():
            print(sweep_key)
            df = spike_obj.df_features  # Extract the features DataFrame from the Spike object
            df['Sweep_#'] = sweep_key  # Add the sweep key as a column
            mega_df = pd.concat([mega_df, df], axis=0)  # Concatenate the DataFrame for each sweep

        # Update columns with file-specific information and check if it exists
        exists = update_columns_at_index(mega_df, file, index)
        if exists:
            super_mega_df = pd.concat([super_mega_df, mega_df], axis=0)  # Concatenate to the final DataFrame
            index += 1
            print("file added")

    return super_mega_df


def monkey_dict(filepaths):
    """
    Processes a list of NWB files to extract spike data and organizes it into a dictionary.

    This function iterates through a list of file paths, extracts spike data from each file, and stores the spike data 
    in a dictionary. Each key in the dictionary is a combination of the file name and the sweep key. 
    The value is an array of spike data for that sweep.

    Parameters:
    - filepaths (list): A list of file paths to NWB files.

    Returns:
    - dict: A dictionary where the keys are strings combining the file name and sweep key, and the values are arrays 
            of spike data.
    """

    spike_dir = {}  #dictionary of sweeps with spikes
    for file in filepaths:
        i = 0
        print(file)
        data = extract_data(file, plot_data = False) #numpy array of all sweeps in the file
                                  
        fileName = os.path.splitext(os.path.basename(file))[0]     #threshold in mVs
        with h5py.File(file, 'r') as f:
            for sweep_key in f['acquisition'].keys():
                if i < len(data):
                    try:
                        sweep_key_obj = Spike(thresh_amp=0, window_length=(5., 5.), smooth_frac=.01)
                        sweep_key_obj.fit(data[i], 20000, n_jobs=-1, progress=tqdm)
                        if sweep_key_obj.n_spikes is not None:
                            key = f"{fileName} {sweep_key}"                        # Concatenate file name with sweep key
                            spike_dir[key] = sweep_key_obj.spikes
                    except ValueError as e:
                        print(f"Fitting failed for sweep {sweep_key}: {e}")

                i += 1
        
    return spike_dir

def plot_correlation_heatmaps(df, columns_to_plot, metadata_param):
    """
    Plots correlation heatmaps for selected columns in a DataFrame, grouped by a specified metadata parameter.

    This function creates correlation heatmaps for a subset of columns in the DataFrame, with each heatmap representing
    data filtered by a unique value of the specified metadata parameter. It calculates the correlation matrix for each
    subset and visualizes it using Seaborn's heatmap.

    Parameters:
    - df (pd.DataFrame): The DataFrame containing the data to plot.
    - columns_to_plot (list of str): A list of column names in the DataFrame to include in the correlation matrix.
    - metadata_param (str): The column name in the DataFrame by which to group and filter the data.

    Returns:
    None: The function displays the correlation heatmaps.
    """
    # Get unique values of the metadata parameter
    unique_values = df[metadata_param].unique()
    
    # Set the style of the heatmap
    sns.set(style="white")
    
    # Loop through each unique value and create a heatmap
    for value in unique_values:
        # Filter the DataFrame for the current unique value
        filtered_df = df[df[metadata_param] == value]

        # Create a correlation matrix
        correlation_matrix = filtered_df[columns_to_plot].corr()
        
        # Plot the heatmap
        plt.figure(figsize=(12, 8))
        sns.heatmap(correlation_matrix, annot=True, cmap='coolwarm', fmt=".2f")
        plt.title(f'Correlation Heatmap of Selected Features for {metadata_param} = {value}')
        plt.show()


def plot_spikes_by_parameter(allMonkey_df, waveform_dict, parameter):
    """
    Plots spike traces for each unique value of a specified parameter in a DataFrame.

    This function generates plots of spike traces from a dictionary of spike data, filtered by different values of a
    specified parameter in the DataFrame. Each plot represents spike traces from sweeps corresponding to a unique value
    of the parameter. It uses a color map to differentiate between different groups.

    Parameters:
    - parameter (str): The column name in the DataFrame by which to group the data and plot spike traces.

    Returns:
    None: The function displays the spike trace plots.
    """
    df = allMonkey_df
    spike_dict = waveform_dict
    
    parameters = df[parameter].unique()

    sns.set(style="whitegrid")  # Use Seaborn 
    
    for param in parameters:
        # Filter the DataFrame for the current parameter value
        param_data = df[df[parameter] == param]

        plt.figure(figsize=(12, 6))
        plt.title(f'Spike Traces for {param}', fontsize=16)
        plt.xlabel('Time (ms)', fontsize=14)
        plt.ylabel('mVs', fontsize=14)

        used_keys = set()
        color_map = plt.get_cmap('tab10')
        
        for index, row in param_data.iterrows():
            monkey_id = row['file_name']
            sweep_num = row['Sweep_#']
            key = f"{monkey_id} {sweep_num}"
            key2 = f"{monkey_id}.nwb {sweep_num}"
           
            if key in spike_dict and key not in used_keys:
                for spk in spike_dict[key]:
                    plt.plot(spk, alpha=0.5, color=color_map(index % 10))  # Use color map for better distinction
                used_keys.add(key)

            if key2 in spike_dict and key2 not in used_keys:
                for spk in spike_dict[key2]:
                    plt.plot(spk, alpha=0.5, color=color_map(index % 10))  # Use color map for better distinction
                used_keys.add(key2)

        plt.grid(True, linestyle='--', alpha=0.6)
        plt.tight_layout()
        plt.show()


def boxplots_by_Param(allMonkey_df_filt, param):
    """
    Generates box plots for specific columns in a DataFrame, grouped by a specified parameter.

    This function creates box plots for a predefined list of columns, with the data grouped by the values of the specified parameter. 
    The box plots provide a visual summary of the distribution of the data for each group, highlighting the median, quartiles, and potential outliers.

    Parameters:
    - param (str): The column name in the DataFrame by which to group the data.
        Each unique value in this column will form a separate group in the box plots.
    - allMonkey_df_filt (pd.Dataframe): data frame from which the data comes from

    Returns:
    None: The function displays the box plots.
    """
    # Set the style of the plots
    sns.set(style="whitegrid")
    
    # Define the list of columns to plot
    columns_to_plot = ['ramp_amp', 'inflection_time', 'inflection_amp', 'peak_amp', 'peak_width', 
                       'peak_sharpness', 'exp_lambda', 'exp_const', 'isi', 'r_squared_ramp']
    
    # Create box plots for each column controlling for 'brainOrigin'
    for column in columns_to_plot:
        plt.figure(figsize=(10, 6))
        sns.boxplot(x=param, y=column, data=allMonkey_df_filt)
        plt.title(f'Box Plot of {column} by {param}')
        plt.xlabel(param)
        plt.ylabel(column)
        plt.xticks(rotation=45)  # Rotate x-axis labels for better visibility
        plt.show()

def plot_waveform(df, combined_dict_filt, metadata_param = None, median = False, Overlapped = False):
    """
    Plots the average spike waveforms with standard deviation shading.

    This function generates plots of average spike waveforms from a dictionary of spike data. The function can filter
    the data based on a specified metadata parameter and plot the results either separately or overlapped.

    Parameters:
    - df (pd.DataFrame): The DataFrame containing metadata about the spike data.
    - combined_dict_filt (dict): A dictionary where keys are formatted as "{file_name} {Sweep_#}" and values are arrays of spike data.
    - metadata_param (str, optional): The column name in the DataFrame to filter the data by. Each unique value in this column will have its own plot.
    - median (bool, optional): If True, calculates the median spike waveform instead of the mean.
    - Overlapped (bool, optional): If True, plots the average waveforms for all unique values of the metadata parameter on a single plot.
        If False, creates separate plots for each unique value.

    Returns:
    None: The function displays the plots.
    """
    if metadata_param is not None:
        parameters = df[metadata_param].unique()

        if Overlapped: # Create the plot
            plt.figure(figsize=(10, 6))
             # Set the labels and title
            plt.xlabel('Time')
            plt.ylabel('mVs')
            plt.ylim(-80, 55)
            plt.title('Overall Average Spike with Standard Deviation for ' + str(metadata_param))
        for value in parameters:
            # Filter the DataFrame for the current unique value
            param_data = df[df[metadata_param] == value]
            all_spikes = []

            used_keys = set()
            for index, row in param_data.iterrows():
                monkey_id = row['file_name']
                sweep_num = row['Sweep_#']
                key = f"{monkey_id} {sweep_num}"
                key2 = f"{monkey_id}.nwb {sweep_num}"
                             
                if key in combined_dict_filt and key not in used_keys:
                    for spk in combined_dict_filt[key]:
                        all_spikes.extend(combined_dict_filt[key])
                    used_keys.add(key)
    
                if key2 in combined_dict_filt and key2 not in used_keys:
                    for spk in combined_dict_filt[key2]:
                        all_spikes.extend(combined_dict_filt[key2])
                    used_keys.add(key2)
                    
            # Convert the list of all spikes to a numpy array for easier manipulation
            all_spikes_array = np.array(all_spikes)

            if not Overlapped:
                # Create the plot
                plt.figure(figsize=(10, 6))
                # Set the labels and title
                plt.xlabel('Time')
                plt.ylabel('mVs')
                plt.ylim(-75, 55)
                plt.title('Overall Average Spike with Standard Deviation for ' + str(value))
            # Calculate the average and standard deviation across all spike arrays
            if median:    
                mean_spike = np.median(all_spikes_array, axis=0)
            else:
                mean_spike = np.mean(all_spikes_array, axis=0)
            std_spike = np.std(all_spikes_array, axis=0)
                    
            plt.plot(mean_spike, label=f'Average Spike {metadata_param}:{value}')
            plt.fill_between(range(len(mean_spike)), mean_spike - std_spike, mean_spike + std_spike,
                             alpha=0.2, label='±1 Standard Deviation')
        
            if not Overlapped:    #plots individual graphs
                plt.legend()
                plt.show()

        if Overlapped:           #plots overlapped graphs
            plt.legend()
            plt.show()

    else:                       #creates an average spike graph of all spikes
        all_spikes = []

        # Aggregate all spike arrays from all keys
        for key, spike_data in combined_dict_filt.items():
            all_spikes.extend(spike_data)
        
        # Convert the list of all spikes to a numpy array for easier manipulation
        all_spikes_array = np.array(all_spikes)
    
        # Calculate the average and standard deviation across all spike arrays
        mean_spike = np.mean(all_spikes_array, axis=0)
        std_spike = np.std(all_spikes_array, axis=0)
    
        # Create the plot
        plt.figure(figsize=(10, 6))
        plt.plot(mean_spike, label='Average Spike', color='b')
        plt.fill_between(range(len(mean_spike)), mean_spike - std_spike, mean_spike + std_spike, 
                         color='b', alpha=0.2, label='±1 Standard Deviation')
        
        # Set the labels and title
        plt.xlabel('Time')
        plt.ylabel('mVs')
        plt.title('Overall Average Spike with Standard Deviation')
        plt.legend()
    
        # Show the plot
        plt.show()



def get_spike_arrays(df, combined_dict_filt, metadata_param):
    """
    Helper function to give spike arrays, or an array of numpy arrays which represent all spikes
    Usefull for graphing 

    """
    parameters = df[metadata_param].unique()
    all_arrays = []
    for value in parameters:
        # Filter the DataFrame for the current unique value
        param_data = df[df[metadata_param] == value]
        all_spikes = []

        used_keys = set()
        for index, row in param_data.iterrows():
            monkey_id = row['file_name']
            sweep_num = row['Sweep_#']
            key = f"{monkey_id} {sweep_num}"
            key2 = f"{monkey_id}.nwb {sweep_num}"
                            
            if key in combined_dict_filt and key not in used_keys:
                for spk in combined_dict_filt[key]:
                    all_spikes.extend(combined_dict_filt[key])
                used_keys.add(key)

            if key2 in combined_dict_filt and key2 not in used_keys:
                for spk in combined_dict_filt[key2]:
                    all_spikes.extend(combined_dict_filt[key2])
                used_keys.add(key2)
                
        # Convert the list of all spikes to a numpy array for easier manipulation
        all_spikes_array = np.array(all_spikes)
        all_arrays.extend(all_spikes_array)
    return all_arrays